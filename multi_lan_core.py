"""
多语言对比工具 - 核心业务逻辑
不包含任何UI代码，可被CLI和GUI复用
"""
import openpyxl
import xml.etree.ElementTree as ET
import html
import difflib
import re
import os
import glob
from typing import Dict, List, Tuple, Optional


class MultiLanguageCore:
    """多语言对比核心业务逻辑"""

    def __init__(self, res_base_path: str = "res"):
        """
        初始化
        :param res_base_path: res目录路径
        """
        self.res_base_path = res_base_path
        self.wb = None
        self.sheet = None
        self.col_a_values = []
        self.countries = []
        self.all_values_from_xml = {}

    def load_excel(self, excel_path: str, sheet_name: str) -> bool:
        """
        加载Excel文件
        :param excel_path: Excel文件路径
        :param sheet_name: Sheet名称
        :return: 是否成功
        """
        try:
            self.wb = openpyxl.load_workbook(excel_path, data_only=False)
            self.sheet = self.wb[sheet_name]
            return True
        except Exception as e:
            raise Exception(f"加载Excel失败: {e}")

    def get_excel_sheets(self, excel_path: str) -> List[str]:
        """
        获取Excel文件的所有Sheet名称
        :param excel_path: Excel文件路径
        :return: Sheet名称列表
        """
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            return wb.sheetnames
        except Exception as e:
            raise Exception(f"读取Excel失败: {e}")

    def get_keys_from_excel(self) -> List[str]:
        """
        从Excel的A列获取所有key值（转小写）
        :return: key列表
        """
        if not self.sheet:
            raise Exception("请先加载Excel文件")

        try:
            self.col_a_values = [
                cell.value.lower() 
                for cell in self.sheet['A'] 
                if cell.value is not None
            ][1:]  # 跳过标题行

            if not self.col_a_values:
                raise Exception("A列没有找到key值")

            return self.col_a_values
        except Exception as e:
            raise Exception(f"获取key值失败: {e}")

    def get_countries_from_excel(self) -> List[str]:
        """
        从Excel第一行获取国家列表
        :return: 国家代码列表
        """
        if not self.sheet:
            raise Exception("请先加载Excel文件")

        try:
            row1_values = list(self.sheet.values)[0]
            self.countries = [val for val in row1_values[1:] if val]

            if not self.countries:
                raise Exception("第一行没有找到国家代码")

            return self.countries
        except Exception as e:
            raise Exception(f"获取国家列表失败: {e}")

    def clean_value(self, value) -> str:
        """
        通用清理函数，处理HTML实体、CDATA、换行符等
        :param value: 原始值
        :return: 清理后的值
        """
        if value is None:
            return ""

        value = str(value)

        # 移除CDATA标签
        value = re.sub(r'<!\[CDATA\[(.*?)\]\]>', r'\1', value, flags=re.DOTALL)

        # HTML实体解码
        value = html.unescape(value)

        # 处理换行符
        value = value.replace('\\n', '\n')
        value = value.replace('\r\n', '\n')
        value = value.replace('\r', '\n')

        # 去除反斜杠
        value = value.replace('\\', '')

        # 去除引号和首尾空格
        value = value.strip('"').strip()

        # 处理每一行，去除多余空格
        value = ' '.join(line.strip() for line in value.split('\n'))

        # 去除连续的空行
        value = ' '.join(line for line in value.split('\n') if line.strip())

        return value

    def read_strings_from_xml(self) -> Tuple[Dict[str, Dict[str, str]], List[str], Dict[str, List[str]]]:
        """
        从XML文件读取字符串资源
        :return: (
            {country: {key: value}} 字典,
            缺失的语言目录列表,
            {country: [missing_keys]} 每个语言中缺失的key
        )
        """
        if not self.countries:
            raise Exception("请先获取国家列表")

        if not self.col_a_values:
            raise Exception("请先获取key值列表")

        self.all_values_from_xml = {}
        missing_xml_files = []  # 找不到的语言目录
        missing_keys_in_xml = {}  # 每个语言中缺失的key

        for country in self.countries:
            # 确定XML文件路径
            if country.lower() in ["values", "value"]:
                file_path = os.path.join(self.res_base_path, "values", "strings.xml")
            else:
                file_path = os.path.join(self.res_base_path, f"values-{country}", "strings.xml")

            # 检查文件是否存在
            if not os.path.exists(file_path):
                missing_xml_files.append(country)
                self.all_values_from_xml[country] = {}
                continue

            try:
                tree = ET.parse(file_path)
                root = tree.getroot()

                key_value = {}
                name_list = []

                for string in root.findall('string'):
                    name = string.get('name')
                    if name:
                        name_list.append(name.lower())

                        if name.lower() in self.col_a_values:
                            # 获取完整文本内容
                            value = ''.join(string.itertext()) if string.text or len(string) > 0 else ""
                            value = self.clean_value(value)
                            key_value[name.lower()] = value

                # 检查缺失的key
                missing_keys = list(set(self.col_a_values) - set(name_list))
                if missing_keys:
                    missing_keys_in_xml[country] = sorted(missing_keys)

                self.all_values_from_xml[country] = key_value

            except ET.ParseError as e:
                raise Exception(f"XML文件解析失败: {file_path}, 错误: {e}")
            except Exception as e:
                raise Exception(f"读取XML文件出错: {file_path}, 错误: {e}")

        return self.all_values_from_xml, missing_xml_files, missing_keys_in_xml

    def get_excel_value_by_key_and_country(self, key: str, country: str) -> str:
        """
        根据key和国家从Excel获取值
        :param key: 键值
        :param country: 国家代码
        :return: Excel中的值
        """
        if country not in self.countries:
            return ""

        country_index = self.countries.index(country) + 2

        for row in self.sheet.iter_rows(
            min_row=2,
            min_col=1,
            max_col=country_index,
            max_row=self.sheet.max_row
        ):
            if row[0].value and row[0].value.lower() == key:
                cell = row[country_index - 1]

                # 处理百分比格式
                if cell.number_format and "%" in cell.number_format:
                    if isinstance(cell.value, (int, float)):
                        if ".0" in cell.number_format:
                            return f"{cell.value:.1%}"
                        else:
                            return f"{int(cell.value * 100)}%"

                # 清理Excel值
                excel_value = self.clean_value(cell.value)
                # 处理HTML标签属性中的双引号
                excel_value = re.sub(r'(\w+)="([^"]*)"', r'\1=\2', excel_value)
                return excel_value

        return ""

    def compare_values(self, value1: str, value2: str) -> bool:
        """
        比较两个值是否完全相同
        :param value1: 值1
        :param value2: 值2
        :return: 是否相同
        """
        v1 = str(value1) if value1 is not None else ""
        v2 = str(value2) if value2 is not None else ""

        ratio = difflib.SequenceMatcher(None, v1, v2).ratio()
        return ratio == 1

    def get_detailed_diff(self, excel_value: str, xml_value: str) -> str:
        """
        获取详细的差异描述
        :param excel_value: Excel中的值
        :param xml_value: XML中的值
        :return: 差异描述
        """
        excel_str = str(excel_value) if excel_value else ""
        xml_str = str(xml_value) if xml_value else ""

        diff_reasons = []

        # 检查长度差异
        if len(excel_str) != len(xml_str):
            diff_reasons.append(f"长度不同(Excel:{len(excel_str)}, XML:{len(xml_str)})")

        # 检查引号差异
        excel_quotes = excel_str.count('"')
        xml_quotes = xml_str.count('"')
        if excel_quotes != xml_quotes:
            diff_reasons.append(f"引号数量不同(Excel:{excel_quotes}, XML:{xml_quotes})")

        # 检查是否只是引号不同
        excel_no_quotes = excel_str.replace('"', '')
        xml_no_quotes = xml_str.replace('"', '')
        if excel_no_quotes == xml_no_quotes:
            diff_reasons.append("仅引号位置/数量不同")

        # 检查HTML标签差异
        if '<' in excel_str or '<' in xml_str:
            if excel_str.replace('"', '') == xml_str.replace('"', ''):
                diff_reasons.append("HTML标签属性引号差异")

        # 检查空格差异
        if excel_str.replace(' ', '') == xml_str.replace(' ', ''):
            diff_reasons.append("仅空格差异")

        # 检查大小写差异
        if excel_str.lower() == xml_str.lower():
            diff_reasons.append("仅大小写差异")

        # 检查特殊字符差异
        excel_clean = re.sub(r'[^\w\s]', '', excel_str)
        xml_clean = re.sub(r'[^\w\s]', '', xml_str)
        if excel_clean == xml_clean:
            diff_reasons.append("仅特殊字符/标点符号差异")

        return "; ".join(diff_reasons) if diff_reasons else "内容完全不同"

    def compare_and_generate_results(self) -> Tuple[int, int]:
        """
        执行对比并生成Excel结果文件
        :return: (差异数量, 相同数量)
        """
        if not self.all_values_from_xml:
            raise Exception("请先读取XML文件")

        # 创建差异结果工作簿
        wb_diff = openpyxl.Workbook()
        ws_diff = wb_diff.active
        ws_diff.title = "对比差异结果"
        ws_diff.append(["Country", "Key", "Excel Value", "XML Value", "差异说明"])

        # 创建相同结果工作簿
        wb_same = openpyxl.Workbook()
        ws_same = wb_same.active
        ws_same.title = "对比相同结果"
        ws_same.append(["Country", "Key", "Excel Value", "XML Value"])

        diff_count = 0
        same_count = 0
        row_num = 2

        # 遍历所有XML值进行对比
        for country, strings in self.all_values_from_xml.items():
            for key, xml_value in strings.items():
                excel_value = self.get_excel_value_by_key_and_country(key, country)

                if not self.compare_values(xml_value, excel_value):
                    # 有差异
                    diff_detail = self.get_detailed_diff(excel_value, xml_value)
                    ws_diff.append([country, key, excel_value, xml_value, diff_detail])
                    diff_count += 1
                else:
                    # 相同
                    ws_same.append([country, key, excel_value, xml_value])
                    same_count += 1
                    row_num += 1

        # 保存结果文件
        wb_diff.save("对比差异结果.xlsx")
        wb_same.save("对比相同结果.xlsx")

        return diff_count, same_count

    @staticmethod
    def find_excel_files() -> List[str]:
        """
        在当前目录查找所有Excel文件（排除对比结果文件）
        :return: Excel文件路径列表
        """
        return glob.glob("**/[!对比][!~]*.xlsx", recursive=True)

    @staticmethod
    def find_apk_files() -> List[str]:
        """
        在当前目录查找所有APK文件
        :return: APK文件路径列表
        """
        return glob.glob("**/*.apk", recursive=True)