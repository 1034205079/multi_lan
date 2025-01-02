import openpyxl
import xml.etree.ElementTree as ET
import html
import difflib
import time
import glob
from colorama import Fore, Style, init

init(autoreset=True)


class MULTI_LAN:

    def __init__(self):
        """加载原始excel文件"""
        xlsx = None  # 提前初始化 xlsx 变量
        sheet = None  # 提前初始化 sheet 变量

        get_excel = glob.glob("**/[!对比][!~]*.xlsx", recursive=True)  # 递归查找所有xlsx文件
        show_xlsx_dict = {key: value for key, value in enumerate(get_excel, 1)}  # 列表显示序号和文件名
        if not show_xlsx_dict:  # 如果没有找到文件
            print("未找到Excel文件！...即将自动退出！")
            time.sleep(5)
            exit()
        elif len(get_excel) == 1:  # 如果只有一个文件
            xlsx = get_excel[0]
            print(f"{Fore.GREEN}仅一个Excel文件，已自动选择：{xlsx}{Style.RESET_ALL}")
        else:  # 如果有多个文件
            while True:  # 循环输入选择
                try:
                    select_xlsx = input(
                        f"{Fore.YELLOW}{show_xlsx_dict}\n请选择要进行对比的Excel文件序号（1-{len(get_excel)}）：{Style.RESET_ALL}")  # 输入序号
                    xlsx = show_xlsx_dict[int(select_xlsx)]  # 选择的文件
                    print(f"{Fore.CYAN}已选择：{xlsx}{Style.RESET_ALL}")
                    break  # 退出循环
                except (ValueError, KeyError):
                    print(f"{Fore.RED}无效的输入！请输入 1 到 {len(get_excel)} 之间的数字。请重新输入。{Style.RESET_ALL}")
                    time.sleep(1)
                    continue  # 继续循环 输入

        """获取全部的sheet"""
        self.wb = openpyxl.load_workbook(xlsx, data_only=False)  # 加载上面已选择的excel文件
        get_sheets = self.wb.sheetnames  # sheet名称出来是列表
        show_sheets_dict = {key: value for key, value in enumerate(get_sheets, 1)}  # 改成字典展示序号和sheet名
        if len(get_sheets) == 1:  # 如果只有一个sheet
            sheet = get_sheets[0]
            print(f"{Fore.GREEN}仅一个sheet，已自动选择：{sheet}{Style.RESET_ALL}")
        else:  # 如果有多个sheet
            while True:  # 循环输入选择
                try:
                    user_input = input(
                        f"\n{Fore.YELLOW}{show_sheets_dict}\n请选择要进行对比的sheet的序号（1-{len(get_sheets)}）：{Style.RESET_ALL}")  # 输入序号
                    sheet = show_sheets_dict[int(user_input)]  # 获取选择的sheet
                    print(f"{Fore.CYAN}已选择：{sheet}{Style.RESET_ALL}")
                    break  # 退出循环
                except (ValueError, KeyError):
                    print(f"{Fore.RED}无效的输入！请输入 1 到 {len(get_sheets)} 之间的数字。请重新输入。{Style.RESET_ALL}")
                    time.sleep(1)
                    continue  # 继续循环 输入
        self.sheet = self.wb[sheet]  # 加载选择的sheet
        time.sleep(3)

    def get_key_from_origin(self):
        """获取key值"""
        try:
            self.col_a_values = [cell.value.lower() for cell in self.sheet['A'] if cell.value is not None][1:]

            if not self.col_a_values:  # 检查是否为空
                print(f"{Fore.RED}请检查key值！...即将自动退出！{Style.RESET_ALL}")
                time.sleep(5)
                exit()

            print(f"从原始文档获取到key值,并转成小写,共计{len(self.col_a_values)}个\n")
            return self.col_a_values

        except AttributeError:
            print(f"{Fore.RED}您选择的sheet中A2往下值不正确，请检查！...即将自动退出！{Style.RESET_ALL}")
            time.sleep(5)
            exit()

    def get_countries_from_origin(self):
        """获取国家列表"""
        row1_values = list(self.sheet.values)[0]
        self.countries = [val for val in row1_values[1:] if val]

        if not self.countries:  # 检查是否为空
            print(f"{Fore.RED}第一行中没有国家代码，请检查！...即将自动退出！{Style.RESET_ALL}")
            time.sleep(5)
            exit()

        print(f"获取到了国家列表：{self.countries}\n")
        return self.countries

    def clean_value(self, value):
        """通用的清理函数"""
        if isinstance(value, str):
            # 处理Excel中的"\n"文本
            value = value.replace('\\n', '\n')

            # 统一换行符
            value = value.replace('\r\n', '\n')
            value = value.replace('\r', '\n')

            # 去除所有反斜杠，但保留换行符
            value = value.replace('\\', '')

            # 去除引号和首尾空格
            value = value.strip('"').strip()

            # 处理每一行，去除多余空格
            value = ' '.join(line.strip() for line in value.split('\n'))

            # 去除连续的空行
            value = ' '.join(line for line in value.split('\n') if line.strip())

        return value

    def read_strings_from_xml(self):
        """从xml读取字符串"""
        self.all_values_from_xml = {}

        for country in self.countries:
            if country.lower() in ["values", "value"]:
                file_path = "res/values/strings.xml"
            else:
                file_path = f"res/values-{country}/strings.xml"

            try:
                tree = ET.parse(file_path)
                root = tree.getroot()

                key_value = {}
                name_list = []
                for string in root.findall('string'):
                    name = string.get('name')
                    if name:
                        name_list.append(name.lower())

                        if name.lower() in self.col_a_values:
                            value = html.unescape(string.text) if string.text else ""
                            value = self.clean_value(value)
                            key_value[name.lower()] = value

                diff = set(self.col_a_values).difference(set(name_list))
                if diff:
                    print(f"国家{country}中这些key：{diff}, 不存在于xml中")
                else:
                    print(f"国家{country}的key都存在于xml中")

                self.all_values_from_xml[country] = key_value

            except FileNotFoundError:
                print(f"xml文件 {file_path} 不存在！请检查命名或文件")
                continue
            except ET.ParseError:
                print(f"xml文件 {file_path} 解析失败！请检查格式")
                continue
        return self.all_values_from_xml

    def get_excel_value_by_key_and_country(self, key, country):
        """根据key和国家获取Excel中的值"""
        country_index = self.countries.index(country) + 2
        for row in self.sheet.iter_rows(min_row=2, min_col=1, max_col=country_index, max_row=self.sheet.max_row):
            if row[0].value and row[0].value.lower() == key:
                cell = row[country_index - 1]
                if cell.number_format and "%" in cell.number_format:  # 如果是百分比格式
                    if isinstance(cell.value, (int, float)):
                        # 检查原始格式中是否有小数点
                        if ".0" in cell.number_format:
                            return f"{cell.value:.1%}"  # 带小数点格式 50.0%
                        else:
                            return f"{int(cell.value * 100)}%"  # 不带小数点格式 50%
                return self.clean_value(cell.value)
        return None

    def compare(self, value1, value2):
        """直接比较两个已清理的值"""
        # 确保两者都是字符串
        if not isinstance(value1, str):
            value1 = str(value1) if value1 is not None else ""
        if not isinstance(value2, str):
            value2 = str(value2) if value2 is not None else ""

        ratio = difflib.SequenceMatcher(None, value1, value2).ratio()
        return ratio == 1

    def compare_and_write_to_excel(self):
        """比较并写入到excel"""
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = "对比差异结果"

        wb_same = openpyxl.Workbook()
        ws_same = wb_same.active
        ws_same.title = "对比相同结果"

        ws_new.append(["Country", "Key", "Excel Value", "XML Value"])
        ws_same.append(["Country", "Key", "Excel Value", "XML Value", "二次校验结果"])

        row_num = 2
        for country, strings in self.all_values_from_xml.items():
            for key, xml_value in strings.items():
                excel_value = self.get_excel_value_by_key_and_country(key, country)
                if not self.compare(xml_value, excel_value):
                    ws_new.append([country, key, excel_value, xml_value])
                else:
                    ws_same.append([country, key, excel_value, xml_value, f'=D{row_num}=C{row_num}'])
                    row_num += 1

        wb_new.save("对比差异结果.xlsx")
        wb_same.save("对比相同结果.xlsx")
        print("\n差异已写入 对比差异结果.xlsx\n相同的结果已写入 对比相同结果.xlsx \n请核实两个文件")


if __name__ == '__main__':
    wb = MULTI_LAN()
    wb.get_key_from_origin()
    wb.get_countries_from_origin()
    wb.read_strings_from_xml()
    wb.compare_and_write_to_excel()
    input("\nPress Enter to exit...")
