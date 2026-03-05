"""
多语言对比工具 - GUI界面版本
使用multi_lan_core和apk_decompiler，避免代码重复
"""
import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import queue
import subprocess
import openpyxl
from multi_lan_core import MultiLanguageCore
from apk_decompiler import APKDecompiler
from excel_processor import ExcelProcessor

# 尝试导入拖放支持
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False


class MultiLanguageGUI:
    """多语言对比工具 - GUI版本"""

    def __init__(self, root):
        self.root = root
        self.root.title("多语言对比工具")
        self.root.geometry("1100x850")
        self.root.minsize(1000, 800)

        # 配色方案
        self.colors = {
            'bg': '#F5F5F5',
            'card_bg': '#FFFFFF',
            'primary': '#2196F3',
            'success': '#4CAF50',
            'warning': '#FF9800',
            'error': '#F44336',
            'text': '#212121',
            'text_secondary': '#757575',
            'border': '#E0E0E0'
        }

        self.root.configure(bg=self.colors['bg'])

        # 核心业务逻辑和工具
        self.core = MultiLanguageCore()
        self.decompiler = None
        self.excel_processor = ExcelProcessor(log_callback=self.log)

        # UI状态变量
        self.package_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        self.selected_sheet = tk.StringVar()
        self.res_base_path = tk.StringVar(value="res")
        self.sheet_checkboxes = {}  # 存储Sheet复选框 {sheet_name: BooleanVar}
        self.sheet_frame = None  # Sheet列表框架
        self.is_processing = False
        self.log_queue = queue.Queue()

        # 构建UI
        self._build_ui()

        # 启用拖放功能
        self._setup_drag_drop()
        self._update_log()

    def _build_ui(self):
        """构建用户界面"""
        main_frame = tk.Frame(self.root, bg=self.colors['bg'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 标题
        self._build_header(main_frame)

        # 左右分栏
        content_frame = tk.Frame(main_frame, bg=self.colors['bg'])
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 左侧操作面板 - 添加滚动支持
        left_container = tk.Frame(content_frame, bg=self.colors['bg'])
        left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 创建Canvas和Scrollbar
        self.left_canvas = tk.Canvas(left_container, bg=self.colors['bg'], highlightthickness=0)
        left_scrollbar = tk.Scrollbar(left_container, orient=tk.VERTICAL, command=self.left_canvas.yview)

        # 创建可滚动的框架
        self.scrollable_left_frame = tk.Frame(self.left_canvas, bg=self.colors['bg'])

        # 配置Canvas
        self.left_canvas.configure(yscrollcommand=left_scrollbar.set)

        # 布局
        left_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 在Canvas中创建窗口
        self.canvas_frame = self.left_canvas.create_window((0, 0), window=self.scrollable_left_frame, anchor=tk.NW)

        # 绑定配置事件
        self.scrollable_left_frame.bind('<Configure>', self._on_left_frame_configure)
        self.left_canvas.bind('<Configure>', self._on_canvas_configure)

        # 绑定鼠标滚轮
        self.left_canvas.bind_all('<MouseWheel>', self._on_mousewheel)

        # 在可滚动框架中构建内容
        self._build_step1_package(self.scrollable_left_frame)
        self._build_step2_excel(self.scrollable_left_frame)
        self._build_step3_actions(self.scrollable_left_frame)
        self._build_results_panel(self.scrollable_left_frame)

        # 右侧日志面板
        right_frame = tk.Frame(content_frame, bg=self.colors['bg'])
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self._build_log_panel(right_frame)

    def _on_left_frame_configure(self, event):
        """更新Canvas滚动区域"""
        self.left_canvas.configure(scrollregion=self.left_canvas.bbox('all'))

    def _on_canvas_configure(self, event):
        """调整Canvas窗口宽度"""
        self.left_canvas.itemconfig(self.canvas_frame, width=event.width)

    def _on_mousewheel(self, event):
        """处理鼠标滚轮事件"""
        self.left_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _setup_drag_drop(self):
        """设置拖放功能"""
        if not HAS_DND:
            self.log("⚠ 未安装 tkinterdnd2，拖放功能不可用", 'warning')
            self.log("  安装方法: pip install tkinterdnd2", 'info')
            return

        # 为 APK 输入框和整个卡片添加拖放支持
        self.package_entry.drop_target_register(DND_FILES)
        self.package_entry.dnd_bind('<<Drop>>', self._on_drop_package)

        self.package_card.drop_target_register(DND_FILES)
        self.package_card.dnd_bind('<<Drop>>', self._on_drop_package)

        # 为 Excel 输入框和整个卡片添加拖放支持
        self.excel_entry.drop_target_register(DND_FILES)
        self.excel_entry.dnd_bind('<<Drop>>', self._on_drop_excel)

        self.excel_card.drop_target_register(DND_FILES)
        self.excel_card.dnd_bind('<<Drop>>', self._on_drop_excel)

        self.log("✓ 拖放功能已启用（可拖到卡片任意位置）", 'success')

    def _on_drop_package(self, event):
        """处理 APK 文件拖放"""
        file_path = event.data
        # 去除可能的大括号
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]

        if file_path.lower().endswith('.apk'):
            self.package_path.set(file_path)
            self.log(f"✓ 已选择包文件: {os.path.basename(file_path)}", 'success')
        else:
            messagebox.showwarning("文件类型错误", "请拖放 APK 文件！")

    def _on_drop_excel(self, event):
        """处理 Excel 文件拖放"""
        file_path = event.data
        # 去除可能的大括号
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]

        if file_path.lower().endswith(('.xlsx', '.xls')):
            self.excel_path.set(file_path)
            self.log(f"✓ 已选择Excel: {os.path.basename(file_path)}", 'success')
            self.load_excel_sheets(file_path)
        else:
            messagebox.showwarning("文件类型错误", "请拖放 Excel 文件（.xlsx 或 .xls）！")

    def _build_header(self, parent):
        """标题区域"""
        header = tk.Frame(parent, bg=self.colors['bg'])
        header.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            header,
            text="🌍 多语言对比工具",
            font=('Microsoft YaHei UI', 24, 'bold'),
            fg=self.colors['text'],
            bg=self.colors['bg']
        ).pack(side=tk.LEFT)

        tk.Label(
            header,
            text="APK资源 ↔ Excel对比分析",
            font=('Microsoft YaHei UI', 10),
            fg=self.colors['text_secondary'],
            bg=self.colors['bg']
        ).pack(side=tk.LEFT, padx=(10, 0), pady=(8, 0))

    def _build_step1_package(self, parent):
        """步骤1：APK反编译"""
        self.package_card = self._create_card(parent)
        card = self.package_card

        tk.Label(
            card,
            text="步骤 1：APK反编译（可选）",
            font=('Microsoft YaHei UI', 11, 'bold'),
            fg=self.colors['text'],
            bg=self.colors['card_bg']
        ).pack(anchor=tk.W, padx=20, pady=(15, 10))

        # 文件路径
        path_frame = tk.Frame(card, bg=self.colors['card_bg'])
        path_frame.pack(fill=tk.X, padx=20, pady=(0, 10))

        self.package_entry = tk.Entry(
            path_frame,
            textvariable=self.package_path,
            font=('Microsoft YaHei UI', 9),
            bg='#FAFAFA',
            relief=tk.FLAT,
            bd=0
        )
        self.package_entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, ipady=6, padx=(0, 5))

        tk.Button(
            path_frame,
            text="选择文件",
            command=self.browse_package,
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['primary'],
            bg=self.colors['card_bg'],
            relief=tk.FLAT,
            cursor='hand2',
            padx=15,
            pady=6,
            bd=1,
            highlightthickness=1,
            highlightbackground=self.colors['primary']
        ).pack(side=tk.LEFT, padx=(0, 5))

        self.decompile_btn = tk.Button(
            path_frame,
            text="反编译",
            command=self.start_decompile,
            font=('Microsoft YaHei UI', 9, 'bold'),
            fg='white',
            bg=self.colors['primary'],
            activebackground='#1976D2',
            relief=tk.FLAT,
            cursor='hand2',
            padx=15,
            pady=6
        )
        self.decompile_btn.pack(side=tk.LEFT)

        # 提示
        tk.Label(
            card,
            text="💡 支持 APK 格式文件（可拖到此卡片任意位置）",
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['text_secondary'],
            bg=self.colors['card_bg']
        ).pack(anchor=tk.W, padx=20, pady=(0, 5))

        # Res路径
        res_frame = tk.Frame(card, bg=self.colors['card_bg'])
        res_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

        tk.Label(
            res_frame,
            text="Res目录：",
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['text_secondary'],
            bg=self.colors['card_bg']
        ).pack(side=tk.LEFT)

        tk.Entry(
            res_frame,
            textvariable=self.res_base_path,
            font=('Microsoft YaHei UI', 9),
            bg='#FAFAFA',
            relief=tk.FLAT,
            bd=0,
            width=30
        ).pack(side=tk.LEFT, ipady=4, padx=(5, 0))

        tk.Button(
            res_frame,
            text="浏览",
            command=self.browse_folder,
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['primary'],
            bg=self.colors['card_bg'],
            relief=tk.FLAT,
            cursor='hand2',
            padx=12,
            pady=4,
            bd=1,
            highlightthickness=1,
            highlightbackground=self.colors['primary']
        ).pack(side=tk.LEFT, padx=(5, 0))

    def _build_step2_excel(self, parent):
        """步骤2：Excel选择"""
        self.excel_card = self._create_card(parent)
        card = self.excel_card

        tk.Label(
            card,
            text="步骤 2：选择Excel文件和Sheet",
            font=('Microsoft YaHei UI', 11, 'bold'),
            fg=self.colors['text'],
            bg=self.colors['card_bg']
        ).pack(anchor=tk.W, padx=20, pady=(15, 10))

        # Excel路径
        excel_frame = tk.Frame(card, bg=self.colors['card_bg'])
        excel_frame.pack(fill=tk.X, padx=20, pady=(0, 10))

        self.excel_entry = tk.Entry(
            excel_frame,
            textvariable=self.excel_path,
            font=('Microsoft YaHei UI', 9),
            bg='#FAFAFA',
            relief=tk.FLAT,
            bd=0
        )
        self.excel_entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, ipady=6, padx=(0, 5))

        tk.Button(
            excel_frame,
            text="选择Excel",
            command=self.browse_excel,
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['primary'],
            bg=self.colors['card_bg'],
            relief=tk.FLAT,
            cursor='hand2',
            padx=15,
            pady=6,
            bd=1,
            highlightthickness=1,
            highlightbackground=self.colors['primary']
        ).pack(side=tk.LEFT)

        # 提示
        tk.Label(
            card,
            text="💡 支持 .xlsx 和 .xls 格式（可拖到此卡片任意位置）",
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['text_secondary'],
            bg=self.colors['card_bg']
        ).pack(anchor=tk.W, padx=20, pady=(5, 10))

        # Sheet列表容器（动态生成）
        self.sheet_list_container = tk.Frame(card, bg=self.colors['card_bg'])
        self.sheet_list_container.pack(fill=tk.X, padx=20, pady=(0, 15))

    def _build_step3_actions(self, parent):
        """步骤3：操作按钮"""
        card = self._create_card(parent)

        tk.Label(
            card,
            text="步骤 3：执行对比",
            font=('Microsoft YaHei UI', 11, 'bold'),
            fg=self.colors['text'],
            bg=self.colors['card_bg']
        ).pack(anchor=tk.W, padx=20, pady=(15, 10))

        btn_frame = tk.Frame(card, bg=self.colors['card_bg'])
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

        self.compare_btn = tk.Button(
            btn_frame,
            text="🔍 开始对比",
            command=self.start_compare,
            font=('Microsoft YaHei UI', 10, 'bold'),
            fg='white',
            bg=self.colors['success'],
            activebackground='#388E3C',
            relief=tk.FLAT,
            cursor='hand2',
            padx=30,
            pady=10
        )
        self.compare_btn.pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(
            btn_frame,
            text="📂 打开结果",
            command=self.open_results,
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['primary'],
            bg=self.colors['card_bg'],
            relief=tk.FLAT,
            cursor='hand2',
            padx=20,
            pady=8,
            bd=1,
            highlightthickness=1,
            highlightbackground=self.colors['primary']
        ).pack(side=tk.LEFT)

    def _build_results_panel(self, parent):
        """结果统计面板"""
        card = self._create_card(parent)

        tk.Label(
            card,
            text="📊 对比结果",
            font=('Microsoft YaHei UI', 11, 'bold'),
            fg=self.colors['text'],
            bg=self.colors['card_bg']
        ).pack(anchor=tk.W, padx=20, pady=(15, 10))

        stats_frame = tk.Frame(card, bg=self.colors['card_bg'])
        stats_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

        # 差异数量
        diff_frame = tk.Frame(stats_frame, bg='#FFEBEE', relief=tk.FLAT, bd=0)
        diff_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        tk.Label(
            diff_frame,
            text="差异",
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['error'],
            bg='#FFEBEE'
        ).pack(pady=(10, 0))

        self.diff_label = tk.Label(
            diff_frame,
            text="0",
            font=('Microsoft YaHei UI', 24, 'bold'),
            fg=self.colors['error'],
            bg='#FFEBEE'
        )
        self.diff_label.pack(pady=(0, 10))

        # 相同数量
        same_frame = tk.Frame(stats_frame, bg='#E8F5E9', relief=tk.FLAT, bd=0)
        same_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(
            same_frame,
            text="相同",
            font=('Microsoft YaHei UI', 9),
            fg=self.colors['success'],
            bg='#E8F5E9'
        ).pack(pady=(10, 0))

        self.same_label = tk.Label(
            same_frame,
            text="0",
            font=('Microsoft YaHei UI', 24, 'bold'),
            fg=self.colors['success'],
            bg='#E8F5E9'
        )
        self.same_label.pack(pady=(0, 10))

    def _build_log_panel(self, parent):
        """日志面板"""
        card = self._create_card(parent)

        tk.Label(
            card,
            text="📋 日志",
            font=('Microsoft YaHei UI', 11, 'bold'),
            fg=self.colors['text'],
            bg=self.colors['card_bg']
        ).pack(anchor=tk.W, padx=20, pady=(15, 10))

        log_frame = tk.Frame(card, bg=self.colors['card_bg'])
        log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))

        self.log_text = tk.Text(
            log_frame,
            font=('Consolas', 9),
            bg='#FAFAFA',
            fg=self.colors['text'],
            relief=tk.FLAT,
            wrap=tk.WORD,
            height=60
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)

        # 配置日志颜色
        self.log_text.tag_config('info', foreground=self.colors['text'])
        self.log_text.tag_config('success', foreground=self.colors['success'])
        self.log_text.tag_config('warning', foreground=self.colors['warning'])
        self.log_text.tag_config('error', foreground=self.colors['error'])
        self.log_text.tag_config('primary', foreground=self.colors['primary'])

    def _create_card(self, parent):
        """创建卡片容器"""
        card = tk.Frame(
            parent,
            bg=self.colors['card_bg'],
            relief=tk.FLAT,
            bd=0,
            highlightthickness=1,
            highlightbackground=self.colors['border']
        )
        card.pack(fill=tk.X, pady=(0, 15))
        return card

    def log(self, message, level='info'):
        """添加日志"""
        self.log_queue.put((message, level))

    def _update_log(self):
        """更新日志显示"""
        try:
            while True:
                message, level = self.log_queue.get_nowait()
                self.log_text.insert(tk.END, f"{message}\n", level)
                self.log_text.see(tk.END)
        except queue.Empty:
            pass
        self.root.after(100, self._update_log)

    def browse_package(self):
        """选择APK文件"""
        filename = filedialog.askopenfilename(
            title="选择APK文件",
            filetypes=[("APK files", "*.apk"), ("All files", "*.*")]
        )
        if filename:
            self.package_path.set(filename)

    def browse_excel(self):
        """选择Excel文件"""
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.excel_path.set(filename)
            self.load_excel_sheets(filename)

    def browse_folder(self):
        """选择res文件夹"""
        folder = filedialog.askdirectory(
            title="选择res目录",
            initialdir=os.getcwd()
        )
        if folder:
            self.res_base_path.set(folder)
            self.log(f"✓ 已选择res目录: {folder}", 'success')

    def load_excel_sheets(self, excel_path):
        """加载Excel的Sheet列表并显示复选框"""
        try:
            # 清空之前的Sheet列表
            for widget in self.sheet_list_container.winfo_children():
                widget.destroy()
            self.sheet_checkboxes.clear()

            # 获取所有Sheet
            sheets = self.core.get_excel_sheets(excel_path)

            if not sheets:
                self.log("⚠ Excel文件中没有找到Sheet", 'warning')
                return

            # 显示Sheet选择标题
            title_frame = tk.Frame(self.sheet_list_container, bg=self.colors['card_bg'])
            title_frame.pack(fill=tk.X, pady=(0, 5))

            tk.Label(
                title_frame,
                text=f"Sheet列表（共 {len(sheets)} 个）：",
                font=('Microsoft YaHei UI', 9, 'bold'),
                fg=self.colors['text'],
                bg=self.colors['card_bg']
            ).pack(side=tk.LEFT)

            # 全选/取消全选按钮
            btn_frame = tk.Frame(title_frame, bg=self.colors['card_bg'])
            btn_frame.pack(side=tk.RIGHT)

            tk.Button(
                btn_frame,
                text="全选",
                command=self.select_all_sheets,
                font=('Microsoft YaHei UI', 8),
                fg=self.colors['primary'],
                bg=self.colors['card_bg'],
                relief=tk.FLAT,
                cursor='hand2',
                padx=8,
                pady=2
            ).pack(side=tk.LEFT, padx=(0, 5))

            tk.Button(
                btn_frame,
                text="取消全选",
                command=self.deselect_all_sheets,
                font=('Microsoft YaHei UI', 8),
                fg=self.colors['text_secondary'],
                bg=self.colors['card_bg'],
                relief=tk.FLAT,
                cursor='hand2',
                padx=8,
                pady=2
            ).pack(side=tk.LEFT)

            # 创建Sheet复选框列表（带滚动条，限制高度，两列布局）
            # 创建容器框架
            list_container = tk.Frame(self.sheet_list_container, bg='#FAFAFA', relief=tk.FLAT, bd=1)
            list_container.pack(fill=tk.X)

            # 创建Canvas和Scrollbar
            canvas = tk.Canvas(list_container, bg='#FAFAFA', highlightthickness=0, height=0)
            scrollbar = tk.Scrollbar(list_container, orient=tk.VERTICAL, command=canvas.yview)

            # 创建可滚动的框架
            list_frame = tk.Frame(canvas, bg='#FAFAFA')

            # 配置Canvas
            canvas.configure(yscrollcommand=scrollbar.set)

            # 为每个Sheet创建复选框 - 两列布局
            for i, sheet_name in enumerate(sheets):
                var = tk.BooleanVar(value=True)  # 默认全选
                self.sheet_checkboxes[sheet_name] = var

                # 计算行列位置
                row = i // 2
                col = i % 2

                cb = tk.Checkbutton(
                    list_frame,
                    text=sheet_name,
                    variable=var,
                    font=('Microsoft YaHei UI', 9),
                    fg=self.colors['text'],
                    bg='#FAFAFA',
                    activebackground='#FAFAFA',
                    selectcolor='#FAFAFA',
                    cursor='hand2',
                    anchor=tk.W
                )
                cb.grid(row=row, column=col, sticky=tk.W, padx=(10, 20), pady=3)

            # 配置列权重，使两列均匀分布
            list_frame.grid_columnconfigure(0, weight=1)
            list_frame.grid_columnconfigure(1, weight=1)

            # 更新Canvas
            canvas.create_window((0, 0), window=list_frame, anchor=tk.NW)
            list_frame.update_idletasks()

            # 计算高度：每行约30px，最多显示3行（6个Sheet）
            rows_count = (len(sheets) + 1) // 2  # 向上取整
            max_height = min(rows_count * 30, 120)  # 最多90px（约3行）
            canvas.configure(height=max_height)

            # 如果超过3行，显示滚动条
            if rows_count > 3:
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            canvas.configure(scrollregion=canvas.bbox('all'))

            self.log(f"✓ 发现 {len(sheets)} 个Sheet，已默认全选", 'success')

        except Exception as e:
            messagebox.showerror("错误", f"读取Excel失败: {e}")
            import traceback
            self.log(traceback.format_exc(), 'error')

    def select_all_sheets(self):
        """全选所有Sheet"""
        for var in self.sheet_checkboxes.values():
            var.set(True)
        self.log("✓ 已全选所有Sheet", 'info')

    def deselect_all_sheets(self):
        """取消全选Sheet"""
        for var in self.sheet_checkboxes.values():
            var.set(False)
        self.log("✓ 已取消全选", 'info')

    def get_selected_sheets(self):
        """获取勾选的Sheet列表"""
        return [name for name, var in self.sheet_checkboxes.items() if var.get()]

    def start_decompile(self):
        """开始反编译APK"""
        if self.is_processing:
            messagebox.showwarning("警告", "正在处理中...")
            return

        apk_path = self.package_path.get()
        if not apk_path:
            messagebox.showerror("错误", "请选择APK文件")
            return

        self.is_processing = True
        self.decompile_btn.config(state=tk.DISABLED, bg='#BDBDBD')

        thread = threading.Thread(target=self._decompile_thread, args=(apk_path,))
        thread.daemon = True
        thread.start()

    def _auto_process_excel(self, excel_path, selected_sheets):
        """自动处理Excel（格式化），返回处理后的文件路径"""
        try:
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            dir_name = os.path.dirname(excel_path)

            # 读取原始Excel
            wb_original = openpyxl.load_workbook(excel_path)

            # 创建新工作簿
            from openpyxl import Workbook
            new_wb = Workbook()
            new_wb.remove(new_wb.active)  # 删除默认sheet

            processed_count = 0
            needs_processing = False

            for sheet_name in selected_sheets:
                if sheet_name not in wb_original.sheetnames:
                    continue

                # 使用processor处理这个sheet
                ws_original = wb_original[sheet_name]
                all_rows = list(ws_original.iter_rows(values_only=True))

                if not all_rows:
                    continue

                # 调用processor的逻辑处理
                processed_header, processed_data = self._process_sheet_data(all_rows)

                if processed_header:
                    # 检查是否需要处理（比较原始和处理后的数据）
                    if len(all_rows) - 1 != len(processed_data) or len(all_rows[0]) != len(processed_header):
                        needs_processing = True

                    # 创建新sheet
                    new_ws = new_wb.create_sheet(title=sheet_name)
                    new_ws.append(processed_header)
                    for row in processed_data:
                        new_ws.append(row)
                    processed_count += 1

            wb_original.close()

            # 如果需要处理，保存新文件
            if processed_count > 0 and needs_processing:
                output_name = f"{base_name}_已处理.xlsx"
                output_path = os.path.join(dir_name, output_name) if dir_name else output_name
                new_wb.save(output_path)
                new_wb.close()

                self.log(f"✓ Excel已自动格式化: {output_name}", 'success')
                self.log(f"  处理了 {processed_count} 个Sheet", 'info')
                return output_path
            else:
                new_wb.close()
                return None

        except Exception as e:
            self.log(f"⚠ Excel自动处理失败: {e}，使用原文件继续", 'warning')
            return None

    def _process_sheet_data(self, all_rows):
        """处理单个Sheet的数据（Excel processor逻辑）"""
        if not all_rows:
            return None, []

        header_row = list(all_rows[0])
        original_cols = len(header_row)

        # 删除空列
        non_empty_cols = []
        for col_idx in range(len(header_row)):
            col_values = [row[col_idx] if col_idx < len(row) else None for row in all_rows]
            if any(v is not None and str(v).strip() != '' for v in col_values):
                non_empty_cols.append(col_idx)

        filtered_header = [header_row[i] for i in non_empty_cols]
        filtered_rows = []
        for row in all_rows[1:]:
            new_row = [row[i] if i < len(row) else None for i in non_empty_cols]
            filtered_rows.append(new_row)

        # 删除空行
        non_empty_rows = []
        for row in filtered_rows:
            if not all(cell is None or str(cell).strip() == '' for cell in row):
                non_empty_rows.append(row)

        # 处理列名
        new_header = []
        cols_to_delete = []

        for i, col in enumerate(filtered_header):
            # 检查是否需要删除
            if col in ['功能模块', '变更版本', '生效状态', '最后更新时间', '更新人', '父记录', 'description']:
                cols_to_delete.append(i)
                continue

            # 转换列名
            new_name = self.excel_processor._get_language_code(col)
            if i == 0 and (new_name is None or new_name == ''):
                new_name = 'xml_key'
            new_header.append(new_name)

        # 从数据中删除指定列
        if cols_to_delete:
            final_data = []
            for row in non_empty_rows:
                new_row = [row[i] for i in range(len(row)) if i not in cols_to_delete]
                final_data.append(new_row)
        else:
            final_data = non_empty_rows

        return new_header, final_data

    def _decompile_thread(self, apk_path):
        """反编译线程"""
        try:
            self.log("=" * 60, 'primary')
            self.log("开始APK反编译", 'primary')
            self.log("=" * 60, 'primary')

            # 创建APKDecompiler，传递日志回调函数
            self.decompiler = APKDecompiler(log_callback=self.log)
            self.decompiler.apk_path = apk_path

            if self.decompiler.decompile():
                res_dir = self.decompiler.get_res_directory()
                if res_dir:
                    self.core.res_base_path = res_dir
                    self.root.after(0, lambda: self.res_base_path.set(res_dir))
                    self.log(f"\n✓ 反编译成功！res目录: {res_dir}", 'success')
                    self.root.after(0, lambda: messagebox.showinfo(
                        "成功", f"反编译完成！\nres目录: {res_dir}"
                    ))
                else:
                    self.log("✗ 未找到res目录", 'error')
            else:
                self.log("✗ 反编译失败", 'error')

        except Exception as e:
            self.log(f"✗ 反编译错误: {e}", 'error')
            import traceback
            self.log(traceback.format_exc(), 'error')
        finally:
            self.is_processing = False
            self.root.after(0, lambda: self.decompile_btn.config(
                state=tk.NORMAL, bg=self.colors['primary']
            ))

    def start_compare(self):
        """开始对比"""
        if self.is_processing:
            messagebox.showwarning("警告", "正在处理中...")
            return

        if not self.excel_path.get():
            messagebox.showerror("错误", "请选择Excel文件")
            return

        selected_sheets = self.get_selected_sheets()
        if not selected_sheets:
            messagebox.showwarning("警告", "请至少勾选一个Sheet进行对比")
            return

        self.is_processing = True
        self.compare_btn.config(state=tk.DISABLED, bg='#BDBDBD')

        thread = threading.Thread(target=self._compare_thread, args=(selected_sheets,))
        thread.daemon = True
        thread.start()

    def _compare_thread(self, selected_sheets):
        """对比线程（支持多Sheet）"""
        try:
            self.log("=" * 60, 'primary')
            self.log("开始多语言对比", 'primary')
            self.log("=" * 60, 'primary')

            self.log(f"\n将对比 {len(selected_sheets)} 个Sheet", 'info')
            for sheet in selected_sheets:
                self.log(f"  - {sheet}", 'info')

            # 自动处理Excel（格式化）
            excel_path = self.excel_path.get()
            self.log("\n检查Excel格式...", 'primary')

            # 尝试处理Excel，如果格式已符合则跳过
            processed_path = self._auto_process_excel(excel_path, selected_sheets)
            if processed_path and processed_path != excel_path:
                self.log(f"✓ 使用处理后的Excel: {os.path.basename(processed_path)}", 'success')
                excel_path = processed_path
            else:
                self.log("✓ Excel格式符合要求，无需处理", 'success')

            # 更新res路径
            self.core.res_base_path = self.res_base_path.get()

            total_diff = 0
            total_same = 0

            # 创建合并结果工作簿
            import openpyxl
            from openpyxl import Workbook
            wb_diff_all = Workbook()
            wb_diff_all.remove(wb_diff_all.active)
            wb_same_all = Workbook()
            wb_same_all.remove(wb_same_all.active)

            # 对每个Sheet分别对比
            for sheet_idx, sheet_name in enumerate(selected_sheets, 1):
                self.log(f"\n{'=' * 60}", 'primary')
                self.log(f"Sheet {sheet_idx}/{len(selected_sheets)}: {sheet_name}", 'primary')
                self.log(f"{'=' * 60}", 'primary')

                try:
                    # 加载Excel
                    self.log("\n加载Excel文件...", 'primary')
                    self.core.load_excel(excel_path, sheet_name)
                    self.log("✓ Excel加载成功", 'success')

                    # 获取keys
                    keys = self.core.get_keys_from_excel()
                    self.log(f"✓ 加载 {len(keys)} 个Key", 'success')

                    # 获取国家列表
                    countries = self.core.get_countries_from_excel()
                    self.log(f"✓ 国家列表: {', '.join(countries)}", 'success')

                    # 读取XML
                    self.log("\n读取XML文件...", 'primary')
                    xml_data, missing_xml_files, missing_keys_in_xml = self.core.read_strings_from_xml()

                    for country in countries:
                        count = len(xml_data.get(country, {}))
                        self.log(f"  {country}: {count} 项", 'info')

                    # 显示缺失的语言目录
                    if missing_xml_files:
                        self.log(f"\n⚠ 警告: {len(missing_xml_files)} 个语言目录未找到", 'warning')
                        for lang in missing_xml_files:
                            self.log(f"  - {lang}", 'warning')

                    # 显示缺失的keys统计
                    if missing_keys_in_xml:
                        total_missing = sum(len(keys) for keys in missing_keys_in_xml.values())
                        self.log(f"\n📊 统计: {len(missing_keys_in_xml)} 个语言中共缺失 {total_missing} 个key", 'info')

                        # 显示所有语言的缺失key列表
                        self.log(f"\n缺失的key详情:", 'primary')
                        for lang, keys in missing_keys_in_xml.items():
                            keys_str = ', '.join(keys)
                            self.log(f"{lang}:[{keys_str}]", 'warning')

                    # 对比
                    self.log("\n开始对比...", 'primary')
                    diff_count, same_count = self.core.compare_and_generate_results()

                    total_diff += diff_count
                    total_same += same_count

                    self.log(f"✓ Sheet '{sheet_name}' 对比完成: 差异={diff_count}, 相同={same_count}", 'success')

                    # 读取临时结果并添加到合并工作簿
                    if os.path.exists("对比差异结果.xlsx"):
                        wb_temp = openpyxl.load_workbook("对比差异结果.xlsx")
                        ws_temp = wb_temp.active
                        ws_new = wb_diff_all.create_sheet(title=sheet_name)
                        for row in ws_temp.iter_rows(values_only=True):
                            ws_new.append(row)
                        wb_temp.close()

                    if os.path.exists("对比相同结果.xlsx"):
                        wb_temp = openpyxl.load_workbook("对比相同结果.xlsx")
                        ws_temp = wb_temp.active
                        ws_new = wb_same_all.create_sheet(title=sheet_name)
                        for row in ws_temp.iter_rows(values_only=True):
                            ws_new.append(row)
                        wb_temp.close()

                except Exception as e:
                    self.log(f"✗ Sheet '{sheet_name}' 对比失败: {e}", 'error')
                    import traceback
                    self.log(traceback.format_exc(), 'error')
                    continue

            # 保存合并结果
            if len(wb_diff_all.sheetnames) > 0:
                wb_diff_all.save("对比差异结果.xlsx")
                wb_same_all.save("对比相同结果.xlsx")

            wb_diff_all.close()
            wb_same_all.close()

            self.root.after(0, lambda: self.diff_label.config(text=str(total_diff)))
            self.root.after(0, lambda: self.same_label.config(text=str(total_same)))

            self.log("\n" + "=" * 60, 'success')
            self.log(f"✓ 全部对比完成！总计: 差异={total_diff}, 相同={total_same}", 'success')
            self.log("=" * 60, 'success')

            self.root.after(0, lambda: messagebox.showinfo(
                "完成",
                f"对比完成！\n\n处理了 {len(selected_sheets)} 个Sheet\n总计差异: {total_diff}\n总计相同: {total_same}\n\n结果已保存到Excel文件"
            ))

        except Exception as e:
            self.log(f"✗ 对比错误: {e}", 'error')
            import traceback
            self.log(traceback.format_exc(), 'error')
            self.root.after(0, lambda: messagebox.showerror("错误", f"对比失败: {e}"))
        finally:
            self.is_processing = False
            self.root.after(0, lambda: self.compare_btn.config(
                state=tk.NORMAL, bg=self.colors['success']
            ))

    def open_results(self):
        """打开结果文件"""
        if os.path.exists("对比差异结果.xlsx"):
            if sys.platform == 'win32':
                os.startfile("对比差异结果.xlsx")
            elif sys.platform == 'darwin':
                subprocess.run(['open', "对比差异结果.xlsx"])
            else:
                subprocess.run(['xdg-open', "对比差异结果.xlsx"])
        else:
            messagebox.showinfo("提示", "请先执行对比")


def main():
    # 如果有 tkinterdnd2，使用 TkinterDnD.Tk()，否则使用普通 tk.Tk()
    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()

    app = MultiLanguageGUI(root)

    # 居中窗口
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    root.mainloop()


if __name__ == '__main__':
    main()